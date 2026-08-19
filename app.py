import streamlit as st
import pandas as pd
import numpy as np
import re
import unicodedata
from copy import copy
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import plotly.graph_objects as go
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ==============================================================================
# CONFIGURACIÓN
# ==============================================================================
APP_VERSION = "4.2 ARPON · HOTEL QUARTZ"
UMBRAL_TOLERANCIA = 1.0
UMBRAL_FOLIO = 0.01

# Prefijos documentales que sí tratamos como folios.
# Se conservan; NO se eliminan durante la normalización.
PREFIJOS_FOLIO = (
    # Prefijos documentales observados en auxiliares ARPON de Hotel Quartz.
    "NCTA", "NC", "H", "B", "R", "X", "E", "S",
)

REFERENCIAS_VACIAS = {
    "", "N/A", "NA", "N.A.", "SIN REF", "SIN REFERENCIA", "S/R",
    "NO APLICA", "NO APLICA.", "NINGUNA", "-", "—", "–", "0",
}

MESES_ES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "aug": 8, "sep": 9, "sept": 9, "oct": 10,
    "nov": 11, "dic": 12, "dec": 12, "jan": 1, "apr": 4,
}

# ==============================================================================
# 1. UTILIDADES
# ==============================================================================

def quitar_acentos(texto):
    texto = "" if texto is None else str(texto)
    return "".join(
        c for c in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(c)
    )


def texto_norm(texto):
    if pd.isna(texto):
        return ""
    s = quitar_acentos(str(texto)).upper().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def concepto_norm(texto):
    """Normalización conservadora para comparar conceptos entre cuentas."""
    s = texto_norm(texto)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_spanish_date(valor):
    if pd.isna(valor):
        return pd.NaT

    if isinstance(valor, (pd.Timestamp, np.datetime64)):
        return pd.Timestamp(valor)

    # Excel puede entregar la fecha como serial numérico, sobre todo en CSV
    # exportados o libros con formatos poco consistentes.
    if isinstance(valor, (int, float, np.integer, np.floating)):
        n = float(valor)
        if 20000 <= n <= 80000:
            try:
                return pd.Timestamp("1899-12-30") + pd.to_timedelta(n, unit="D")
            except Exception:
                return pd.NaT

    s = str(valor).strip()
    if not s:
        return pd.NaT

    # Serial de Excel guardado como texto.
    if re.fullmatch(r"\d{5}(?:\.\d+)?", s):
        n = float(s)
        if 20000 <= n <= 80000:
            return pd.Timestamp("1899-12-30") + pd.to_timedelta(n, unit="D")

    # dd/Mmm/aaaa, aceptando abreviaturas ES/EN.
    m = re.match(
        r"^(\d{1,2})[/\-]([A-Za-zÁÉÍÓÚáéíóúÑñ]{3,4})[/\-](\d{4})$",
        s,
    )
    if m:
        day, mon_abbr, year = m.groups()
        mon = quitar_acentos(mon_abbr).lower()
        if mon in MESES_ES:
            try:
                return pd.Timestamp(
                    year=int(year), month=MESES_ES[mon], day=int(day)
                )
            except ValueError:
                return pd.NaT

    # ISO yyyy-mm-dd / yyyy-mm-dd hh:mm:ss.
    if re.match(r"^\d{4}-\d{1,2}-\d{1,2}", s):
        return pd.to_datetime(s, errors="coerce")

    # Evita interpretar números aislados (por ejemplo "2026") como fechas.
    if not re.search(r"[/\-]", s) and not re.search(r"[A-Za-zÁÉÍÓÚáéíóúÑñ]", s):
        return pd.NaT

    return pd.to_datetime(s, dayfirst=True, errors="coerce")


def parse_amount(valor):
    """
    Convierte montos contables sin convertir silenciosamente texto inválido en cero.
    Soporta:
      1,234.56
      $1,234.56
      (1,234.56)
      -1,234.56
      vacío -> 0
    """
    if pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float, np.integer, np.floating)):
        if pd.isna(valor):
            return 0.0
        return float(valor)

    s = str(valor).strip()
    if s == "":
        return 0.0

    negativo_parentesis = s.startswith("(") and s.endswith(")")
    if negativo_parentesis:
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "").replace(" ", "")
    s = s.replace("MXN", "").replace("USD", "")

    # Guiones aislados suelen representar vacío.
    if s in {"-", "—", "–"}:
        return 0.0

    try:
        n = float(s)
        return -n if negativo_parentesis else n
    except ValueError:
        return np.nan


def columna_a_monto(serie):
    return serie.apply(parse_amount)


def es_vacio(valor):
    return pd.isna(valor) or str(valor).strip() == ""


def extraer_numero_folio(ref_norm):
    if not ref_norm:
        return None
    m = re.search(r"(\d+)$", str(ref_norm))
    return m.group(1) if m else None


def normalizar_referencia_base(ref):
    """
    Normaliza SIN destruir prefijos documentales.

    Ejemplos:
      NCTA12846    -> NCTA12846
      NCTA-13547   -> NCTA13547
      NCTA 14314   -> NCTA14314
      NC11404      -> NC11404
      FACTURA NCTA-13547 -> NCTA13547
      8729         -> 8729

    Referencias libres se conservan en forma normalizada, pero NO se marcan
    automáticamente como folios.
    """
    if es_vacio(ref):
        return None, "VACIA", None

    # Evitar 123.0 cuando Excel leyó un folio numérico como float.
    if isinstance(ref, float) and ref.is_integer():
        s = str(int(ref))
    else:
        s = str(ref).strip()

    s = texto_norm(s)
    if s in REFERENCIAS_VACIAS:
        return None, "VACIA", None

    s = re.sub(
        r"^(?:FACTURA|FAC|FOLIO|REF|REFERENCIA)\s*[:.\-]?\s*",
        "",
        s,
    )

    # Folio con prefijo conocido.
    prefijos = "|".join(sorted(PREFIJOS_FOLIO, key=len, reverse=True))
    m = re.fullmatch(rf"({prefijos})[\s\-_/.:]*(\d+)", s)
    if m:
        pref, num = m.groups()
        return f"{pref}{num}", "FOLIO_PREFIJO", pref

    # Folio completamente numérico.
    if re.fullmatch(r"\d+", s):
        return s, "FOLIO_NUMERICO", None

    # Referencia libre / bancaria / descriptiva.
    libre = re.sub(r"\s+", " ", s).strip()
    return libre if libre else None, "OTRA_REFERENCIA", None


def extraer_referencia_de_concepto(concepto):
    """
    Recupera un folio del concepto SOLO cuando existe un candidato documental
    inequívoco. No intenta inferir folios a partir de cualquier número.
    """
    s = texto_norm(concepto)
    if not s:
        return None

    prefijos = "|".join(sorted(PREFIJOS_FOLIO, key=len, reverse=True))
    patron = re.compile(
        rf"(?<![A-Z0-9])({prefijos})[\s\-_/.:]*(\d{{2,}})(?!\d)"
    )
    encontrados = {
        f"{m.group(1)}{m.group(2)}" for m in patron.finditer(s)
    }
    if len(encontrados) == 1:
        return next(iter(encontrados))
    return None


def enriquecer_referencias(df):
    df = df.copy()
    df["referencia_original"] = df["referencia"]

    refs_norm = []
    refs_tipo = []
    refs_prefijo = []
    refs_fuente = []
    refs_recuperadas = []

    for _, row in df.iterrows():
        original = row["referencia_original"]
        norm, tipo, prefijo = normalizar_referencia_base(original)
        recuperada = False
        fuente = "original"

        if norm is None:
            rec = extraer_referencia_de_concepto(row.get("concepto", ""))
            if rec:
                norm, tipo, prefijo = normalizar_referencia_base(rec)
                fuente = "concepto"
                recuperada = True
            else:
                fuente = "vacia"

        refs_norm.append(norm)
        refs_tipo.append(tipo)
        refs_prefijo.append(prefijo)
        refs_fuente.append(fuente)
        refs_recuperadas.append(recuperada)

    df["referencia_norm"] = refs_norm
    df["referencia_tipo"] = refs_tipo
    df["referencia_prefijo"] = refs_prefijo
    df["referencia_fuente"] = refs_fuente
    df["referencia_recuperada"] = refs_recuperadas
    df["referencia_numero"] = df["referencia_norm"].apply(extraer_numero_folio)

    df["tiene_referencia"] = df["referencia_norm"].notna()
    df["es_folio"] = df["referencia_tipo"].isin(
        ["FOLIO_PREFIJO", "FOLIO_NUMERICO"]
    )
    return df


def cargar_archivo_robusto(file_bytes, file_name):
    """
    Lee Excel o CSV. CSV intenta UTF-8 antes de latin-1 para evitar mojibake.
    """
    bio = BytesIO(file_bytes)
    lower = file_name.lower()

    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        return pd.read_excel(bio, header=None)

    # Incluso si la extensión es CSV, primero intentamos Excel por seguridad
    # cuando el contenido realmente lo sea.
    try:
        return pd.read_excel(bio, header=None)
    except Exception:
        pass

    errores = []
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            bio.seek(0)
            return pd.read_csv(
                bio,
                header=None,
                encoding=encoding,
                sep=None,
                engine="python",
            )
        except Exception as e:
            errores.append(f"{encoding}: {e}")

    raise ValueError(
        "No fue posible leer el archivo como Excel ni CSV. "
        + " | ".join(errores[-2:])
    )


def to_excel_workbook(tablas):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        usados = set()
        for nombre, df in tablas.items():
            hoja = re.sub(r"[\[\]\*\?/\\:]", "_", str(nombre))[:31] or "Datos"
            base = hoja
            i = 2
            while hoja in usados:
                suf = f"_{i}"
                hoja = (base[: 31 - len(suf)] + suf)
                i += 1
            usados.add(hoja)
            (df if df is not None else pd.DataFrame()).to_excel(
                writer, sheet_name=hoja, index=False
            )
    return output.getvalue()


def construir_empresa_uid(sistema_origen, empresa):
    """Identidad estable para impedir cruces entre sistemas o empresas distintas."""
    empresa_norm = texto_norm(empresa)
    if not empresa_norm:
        empresa_norm = "SIN_EMPRESA_IDENTIFICADA"
    return f"{sistema_origen}::{empresa_norm}"


def agregar_identidad_origen(df, sistema_origen, empresa, file_name):
    """Agrega identidad de sistema, empresa y cuenta a movimientos/resúmenes."""
    x = df.copy()
    empresa_txt = "" if empresa is None else str(empresa).strip()
    empresa_uid = construir_empresa_uid(sistema_origen, empresa_txt)
    x["sistema_origen"] = sistema_origen
    x["empresa"] = empresa_txt
    x["empresa_uid"] = empresa_uid
    x["cuenta_logica_uid"] = (
        empresa_uid + "::" + x["meta_codigo"].astype(str)
    )
    # cuenta_uid distingue la ocurrencia del mismo código dentro de cada archivo.
    x["cuenta_uid"] = (
        empresa_uid + "::" + str(file_name) + "::" + x["meta_codigo"].astype(str)
    )
    return x


# ==============================================================================
# 2. MOTOR ARPON DE LECTURA Y VALIDACIÓN
# ==============================================================================

def validar_formato_arpon(raw, file_name):
    """Valida que el archivo tenga la firma del Auxiliar de Cuentas de ARPON."""
    if raw.empty:
        raise ValueError(f"{file_name}: el archivo está vacío.")

    if raw.shape[1] < 7:
        raise ValueError(
            f"{file_name}: este auditor es exclusivo para ARPON y espera al menos "
            "7 columnas: Póliza | Fecha | Docto. | Concepto | Cargo | Abono | Saldo."
        )

    for _, row in raw.iterrows():
        vals = [texto_norm(row.iloc[i]) for i in range(7)]
        if (
            vals[0] == "POLIZA"
            and vals[1] == "FECHA"
            and vals[2].startswith("DOCTO")
            and vals[3] == "CONCEPTO"
            and vals[4] == "CARGO"
            and vals[5] == "ABONO"
            and vals[6] == "SALDO"
        ):
            return True

    primeras = raw.head(10).fillna("").astype(str).to_string(index=False, header=False)
    raise ValueError(
        f"{file_name}: el archivo no corresponde al Auxiliar de Cuentas de ARPON. "
        "Este auditor solo acepta exportaciones ARPON. Se espera la estructura "
        "Póliza | Fecha | Docto. | Concepto | Cargo | Abono | Saldo. "
        f"Primeras filas detectadas:\n{primeras[:800]}"
    )

def extraer_empresa_periodo_arpon(raw):
    """Extrae empresa y periodo del Auxiliar de Cuentas exportado por ARPON."""
    empresa = None
    periodo_inicio = pd.NaT
    periodo_fin = pd.NaT

    idx_aux = None
    texto_aux = None
    for idx in raw.index:
        for col in range(min(raw.shape[1], 4)):
            t = texto_norm(raw.iloc[idx, col])
            if "AUXILIAR DE CUENTAS" in t:
                idx_aux = idx
                texto_aux = str(raw.iloc[idx, col])
                break
        if idx_aux is not None:
            break

    if idx_aux is not None:
        for j in range(idx_aux - 1, -1, -1):
            candidato = raw.iloc[j, 0] if raw.shape[1] else None
            if es_vacio(candidato):
                continue
            c = str(candidato).strip()
            cn = texto_norm(c)
            if cn.startswith("MON ") or cn.startswith("TUE ") or cn.startswith("WED "):
                continue
            if cn.startswith("THU ") or cn.startswith("FRI ") or cn.startswith("SAT "):
                continue
            if cn.startswith("SUN "):
                continue
            empresa = c
            break

    if texto_aux:
        m = re.search(
            r"DEL\s+(.+?)\s+AL\s+(.+?)$",
            texto_norm(texto_aux),
            flags=re.I,
        )
        if m:
            periodo_inicio = parse_spanish_date(m.group(1).title())
            periodo_fin = parse_spanish_date(m.group(2).title())

    return empresa, periodo_inicio, periodo_fin


def validar_secuencia_saldo(movs, resumen):
    """
    Valida el saldo acumulado movimiento por movimiento usando las dos
    naturalezas posibles. Devuelve diagnóstico por cuenta.
    """
    resultados = []
    mapa_si = resumen.set_index("cuenta_uid")["saldo_inicial"]

    for cuenta_uid, mm in movs.groupby("cuenta_uid"):
        mm = mm.sort_values("fila_origen").copy()
        saldo_ini = float(mapa_si.loc[cuenta_uid])

        prev = mm["saldo_acumulado"].shift(1)
        prev.iloc[0] = saldo_ini

        esperado_deud = prev + mm["cargos"] - mm["abonos"]
        esperado_acre = prev - mm["cargos"] + mm["abonos"]
        err_deud = (mm["saldo_acumulado"] - esperado_deud).abs()
        err_acre = (mm["saldo_acumulado"] - esperado_acre).abs()

        max_deud = float(err_deud.max()) if len(err_deud) else 0.0
        max_acre = float(err_acre.max()) if len(err_acre) else 0.0

        if max_deud <= UMBRAL_TOLERANCIA and max_acre > UMBRAL_TOLERANCIA:
            nat = "DEUDORA"
            err_elegido = err_deud
        elif max_acre <= UMBRAL_TOLERANCIA and max_deud > UMBRAL_TOLERANCIA:
            nat = "ACREEDORA"
            err_elegido = err_acre
        elif max_deud <= UMBRAL_TOLERANCIA and max_acre <= UMBRAL_TOLERANCIA:
            nat = "INDETERMINADA"
            err_elegido = pd.concat([err_deud, err_acre], axis=1).min(axis=1)
        elif max_deud < max_acre:
            nat = "DEUDORA"
            err_elegido = err_deud
        else:
            nat = "ACREEDORA"
            err_elegido = err_acre

        resultados.append(
            {
                "cuenta_uid": cuenta_uid,
                "naturaleza_secuencia": nat,
                "n_errores_saldo_secuencia": int(
                    (err_elegido > UMBRAL_TOLERANCIA).sum()
                ),
                "max_error_saldo_secuencia": float(err_elegido.max()),
                "max_error_deudora_secuencia": max_deud,
                "max_error_acreedora_secuencia": max_acre,
                "ultimo_saldo_movimiento": float(mm.iloc[-1]["saldo_acumulado"]),
            }
        )

    return pd.DataFrame(resultados)


def procesar_formato_arpon(raw, file_name):
    """
    Procesa el Auxiliar de Cuentas de ARPON usado por Hotel Quartz:
      Cuenta: código - nombre                                  saldo inicial
      Póliza | Fecha | Docto. | Concepto | Cargo | Abono | Saldo
      ...
             Totales                         cargos | abonos | saldo final
             Neto Periodo                           | neto
    """
    if raw.shape[1] < 7:
        raise ValueError(
            f"{file_name}: el auxiliar ARPON requiere al menos 7 columnas."
        )

    empresa, periodo_inicio, periodo_fin = extraer_empresa_periodo_arpon(raw)

    # Encabezados de cuenta.
    patron_header = re.compile(
        r"^CUENTA:\s*(\d+(?:-\d+){2,})\s*-\s*(.+)$",
        flags=re.I,
    )
    headers = []
    for idx in raw.index:
        t = texto_norm(raw.iloc[idx, 0])
        m = patron_header.match(t)
        if not m:
            continue
        saldo_ini = parse_amount(raw.iloc[idx, 6])
        if pd.isna(saldo_ini):
            raise ValueError(
                f"{file_name}: no pude leer el saldo inicial de la cuenta "
                f"{m.group(1)} en la fila Excel {idx + 1}."
            )
        headers.append(
            {
                "idx": idx,
                "codigo": m.group(1),
                "nombre": m.group(2).strip(),
                "saldo_inicial": float(saldo_ini),
            }
        )

    if not headers:
        raise ValueError(
            f"{file_name}: no se detectó ninguna fila 'Cuenta: código - nombre'."
        )

    df = raw.copy()
    df["meta_codigo"] = pd.Series(index=df.index, dtype="object")
    df["meta_nombre"] = pd.Series(index=df.index, dtype="object")
    df["meta_saldo_inicial"] = pd.Series(index=df.index, dtype="float64")

    for h in headers:
        df.loc[h["idx"], "meta_codigo"] = h["codigo"]
        df.loc[h["idx"], "meta_nombre"] = h["nombre"]
        df.loc[h["idx"], "meta_saldo_inicial"] = h["saldo_inicial"]

    df["meta_codigo"] = df["meta_codigo"].ffill()
    df["meta_nombre"] = df["meta_nombre"].ffill()
    df["meta_saldo_inicial"] = df["meta_saldo_inicial"].ffill()

    # Movimiento = fecha válida en columna B + póliza en A + cuenta ya activa.
    fechas_candidato = raw[1].apply(parse_spanish_date)
    etiquetas_b = raw[1].apply(texto_norm)
    is_mov = (
        fechas_candidato.notna()
        & raw[0].apply(lambda x: not es_vacio(x))
        & df["meta_codigo"].notna()
        & ~etiquetas_b.isin({"TOTALES", "NETO PERIODO"})
    )

    if not is_mov.any():
        raise ValueError(f"{file_name}: no se detectaron movimientos válidos.")

    movs = df[is_mov].copy()
    movs = movs.rename(
        columns={
            0: "poliza",
            1: "fecha_raw",
            2: "referencia",
            3: "concepto",
            4: "cargos",
            5: "abonos",
            6: "saldo_acumulado",
        }
    )
    movs["tipo_poliza"] = (
        movs["poliza"].astype(str).str.extract(r"^([A-Za-z]+)", expand=False)
        .fillna("").str.upper()
    )
    movs["fila_origen"] = movs.index + 1
    movs["archivo"] = file_name
    movs["periodo_inicio"] = periodo_inicio
    movs["periodo_fin"] = periodo_fin
    movs = agregar_identidad_origen(movs, "ARPON", empresa or "", file_name)
    movs["fecha"] = movs["fecha_raw"].apply(parse_spanish_date)

    for c in ["cargos", "abonos", "saldo_acumulado"]:
        original = movs[c].copy()
        convertido = columna_a_monto(original)
        invalidos = convertido.isna()
        if invalidos.any():
            filas = movs.loc[invalidos, "fila_origen"].tolist()
            ejemplos = original[invalidos].astype(str).head(5).tolist()
            raise ValueError(
                f"{file_name}: valores no numéricos en '{c}' en "
                f"{len(filas)} movimiento(s). Filas {filas[:10]}; "
                f"ejemplos: {ejemplos}."
            )
        movs[c] = convertido.astype(float)

    movs["concepto_norm"] = movs["concepto"].apply(concepto_norm)
    movs = enriquecer_referencias(movs)

    # Resumen base por cuenta a partir del detalle.
    resumen_rows = []
    for h in headers:
        uid = f"{construir_empresa_uid('ARPON', empresa or '')}::{file_name}::{h['codigo']}"
        mm = movs[movs["cuenta_uid"] == uid].sort_values("fila_origen")
        if mm.empty:
            total_cargos = 0.0
            total_abonos = 0.0
            saldo_final = h["saldo_inicial"]
        else:
            total_cargos = float(mm["cargos"].sum())
            total_abonos = float(mm["abonos"].sum())
            saldo_final = float(mm.iloc[-1]["saldo_acumulado"])

        resumen_rows.append(
            {
                "archivo": file_name,
                "sistema_origen": "ARPON",
                "empresa": empresa or "",
                "empresa_uid": construir_empresa_uid("ARPON", empresa or ""),
                "cuenta_logica_uid": f"{construir_empresa_uid('ARPON', empresa or '')}::{h['codigo']}",
                "periodo_inicio": periodo_inicio,
                "periodo_fin": periodo_fin,
                "cuenta_uid": uid,
                "meta_codigo": h["codigo"],
                "meta_nombre": h["nombre"],
                "saldo_inicial": h["saldo_inicial"],
                "total_cargos": total_cargos,
                "total_abonos": total_abonos,
                "saldo_final_aux": saldo_final,
            }
        )

    resumen = pd.DataFrame(resumen_rows)

    # Totales explícitos del reporte.
    mask_totales = raw[1].apply(texto_norm).eq("TOTALES")
    total_rows = raw.index[mask_totales].tolist()
    n_totales = len(total_rows)
    gran_total = None
    amarre_totales = None

    if n_totales:
        # Si hay exactamente un total y una cuenta, úsalo como total certificado.
        if n_totales == 1 and len(resumen) == 1:
            idx = total_rows[0]
            tc = parse_amount(raw.iloc[idx, 4])
            ta = parse_amount(raw.iloc[idx, 5])
            sf = parse_amount(raw.iloc[idx, 6])
            if any(pd.isna(x) for x in [tc, ta, sf]):
                raise ValueError(
                    f"{file_name}: no fue posible leer la fila Totales "
                    f"(fila Excel {idx + 1})."
                )
            resumen.loc[0, "total_cargos"] = float(tc)
            resumen.loc[0, "total_abonos"] = float(ta)
            resumen.loc[0, "saldo_final_aux"] = float(sf)
            gran_total = float(sf)
            amarre_totales = (
                abs(float(tc) - float(movs["cargos"].sum())) <= UMBRAL_TOLERANCIA
                and abs(float(ta) - float(movs["abonos"].sum())) <= UMBRAL_TOLERANCIA
            )
        elif n_totales == len(resumen):
            # Posible total por cada cuenta: la cuenta activa se obtiene por ffill.
            explicitos = []
            for idx in total_rows:
                codigo = df.loc[idx, "meta_codigo"]
                if pd.isna(codigo):
                    continue
                explicitos.append(
                    {
                        "meta_codigo": str(codigo),
                        "total_cargos_exp": parse_amount(raw.iloc[idx, 4]),
                        "total_abonos_exp": parse_amount(raw.iloc[idx, 5]),
                        "saldo_final_exp": parse_amount(raw.iloc[idx, 6]),
                    }
                )
            exp = pd.DataFrame(explicitos)
            if len(exp) == len(resumen) and exp["meta_codigo"].nunique() == len(resumen):
                resumen = resumen.merge(exp, on="meta_codigo", how="left")
                for dst, src in [
                    ("total_cargos", "total_cargos_exp"),
                    ("total_abonos", "total_abonos_exp"),
                    ("saldo_final_aux", "saldo_final_exp"),
                ]:
                    resumen[dst] = resumen[src].astype(float)
                resumen = resumen.drop(
                    columns=["total_cargos_exp", "total_abonos_exp", "saldo_final_exp"]
                )
                gran_total = float(resumen["saldo_final_aux"].sum())
                amarre_totales = True

    # Comparar detalle contra los totales de cada cuenta.
    sum_mov = (
        movs.groupby("cuenta_uid", as_index=False)
        .agg(mov_cargos=("cargos", "sum"), mov_abonos=("abonos", "sum"))
    )
    resumen = resumen.merge(sum_mov, on="cuenta_uid", how="left")
    resumen[["mov_cargos", "mov_abonos"]] = resumen[
        ["mov_cargos", "mov_abonos"]
    ].fillna(0.0)
    resumen["dif_cargos_vs_total"] = resumen["total_cargos"] - resumen["mov_cargos"]
    resumen["dif_abonos_vs_total"] = resumen["total_abonos"] - resumen["mov_abonos"]

    mal_detalle = resumen[
        (resumen["dif_cargos_vs_total"].abs() > UMBRAL_TOLERANCIA)
        | (resumen["dif_abonos_vs_total"].abs() > UMBRAL_TOLERANCIA)
    ]
    if not mal_detalle.empty:
        raise ValueError(
            f"{file_name}: el detalle no amarra con la fila Totales del "
            "auxiliar ARPON."
        )

    # Validación independiente: secuencia completa del saldo acumulado.
    sec = validar_secuencia_saldo(movs, resumen)
    resumen = resumen.merge(sec, on="cuenta_uid", how="left")
    errores_secuencia = int(resumen["n_errores_saldo_secuencia"].fillna(0).sum())
    max_error_secuencia = float(
        resumen["max_error_saldo_secuencia"].fillna(0).max()
    )
    if errores_secuencia:
        raise ValueError(
            f"{file_name}: se detectaron {errores_secuencia} movimiento(s) cuya "
            "secuencia de saldo acumulado no puede reproducirse con cargos/abonos."
        )

    # Neto del periodo, si está impreso.
    mask_neto = raw[1].apply(texto_norm).eq("NETO PERIODO")
    neto_periodo = None
    amarre_neto = None
    if mask_neto.any():
        idx_neto = raw.index[mask_neto][-1]
        candidatos = [raw.iloc[idx_neto, c] for c in range(4, min(7, raw.shape[1]))]
        for valor in candidatos:
            if es_vacio(valor):
                continue
            n = parse_amount(valor)
            if not pd.isna(n):
                neto_periodo = float(n)
                break
        if neto_periodo is not None:
            neto_calc = float(movs["cargos"].sum() - movs["abonos"].sum())
            amarre_neto = abs(abs(neto_calc) - abs(neto_periodo)) <= UMBRAL_TOLERANCIA

    if gran_total is None:
        gran_total = float(resumen["saldo_final_aux"].sum())

    suma_saldos = float(resumen["saldo_final_aux"].sum())
    amarre_gran_total = abs(suma_saldos - gran_total) <= max(
        UMBRAL_TOLERANCIA, abs(gran_total) * 1e-6
    )

    diag = {
        "archivo": file_name,
        "sistema_origen": "ARPON",
        "formato": "ARPON_AUXILIAR_CUENTAS",
        "empresa": empresa or "",
        "periodo_inicio": periodo_inicio,
        "periodo_fin": periodo_fin,
        "n_headers": len(headers),
        "n_totales": n_totales,
        "n_movs": int(len(movs)),
        "gran_total": gran_total,
        "suma_saldos_cuenta": suma_saldos,
        "n_candidatos_no_header": 0,
        "amarre_gran_total": amarre_gran_total,
        "amarre_totales_detalle": amarre_totales,
        "neto_periodo": neto_periodo,
        "amarre_neto_periodo": amarre_neto,
        "n_errores_saldo_secuencia": errores_secuencia,
        "max_error_saldo_secuencia": max_error_secuencia,
    }

    return movs.reset_index(drop=True), resumen.reset_index(drop=True), diag


def procesar_archivo_core(file_bytes, file_name):
    raw = cargar_archivo_robusto(file_bytes, file_name)
    validar_formato_arpon(raw, file_name)
    return procesar_formato_arpon(raw, file_name)


@st.cache_data(show_spinner=False)
def procesar_archivo_engine(file_bytes, file_name):
    return procesar_archivo_core(file_bytes, file_name)

# ==============================================================================
# 3. NATURALEZA CONTABLE Y CONCILIACIÓN ARPON
# ==============================================================================

def detectar_naturaleza(resumen, movs):
    """
    Detecta naturaleza por cuenta comparando las dos ecuaciones posibles
    contra el saldo final reportado por ARPON.

    Deudora:
      final = inicial + cargos - abonos

    Acreedora:
      final = inicial - cargos + abonos

    También usa el comportamiento del saldo acumulado como evidencia secundaria.
    """
    r = resumen.copy()

    r["esperado_deudora"] = (
        r["saldo_inicial"] + r["total_cargos"] - r["total_abonos"]
    )
    r["esperado_acreedora"] = (
        r["saldo_inicial"] - r["total_cargos"] + r["total_abonos"]
    )
    r["error_deudora"] = r["saldo_final_aux"] - r["esperado_deudora"]
    r["error_acreedora"] = r["saldo_final_aux"] - r["esperado_acreedora"]

    naturalezas = []
    confianzas = []

    for _, row in r.iterrows():
        ed = abs(row["error_deudora"])
        ea = abs(row["error_acreedora"])

        deud_cuadra = ed <= UMBRAL_TOLERANCIA
        acre_cuadra = ea <= UMBRAL_TOLERANCIA

        if deud_cuadra and not acre_cuadra:
            naturaleza = "DEUDORA"
            confianza = "ALTA"
        elif acre_cuadra and not deud_cuadra:
            naturaleza = "ACREEDORA"
            confianza = "ALTA"
        elif deud_cuadra and acre_cuadra:
            # Puede ocurrir si cargos == abonos. Revisamos el saldo acumulado.
            mm = movs[movs["cuenta_uid"] == row["cuenta_uid"]].copy()
            if len(mm):
                mm = mm.sort_values("fila_origen")
                saldo_previo = mm["saldo_acumulado"].shift(1)
                saldo_previo.iloc[0] = row["saldo_inicial"]
                delta_obs = mm["saldo_acumulado"] - saldo_previo

                err_mov_deud = (
                    delta_obs - (mm["cargos"] - mm["abonos"])
                ).abs().sum()
                err_mov_acre = (
                    delta_obs - (mm["abonos"] - mm["cargos"])
                ).abs().sum()

                if err_mov_deud + 0.01 < err_mov_acre:
                    naturaleza = "DEUDORA"
                    confianza = "ALTA"
                elif err_mov_acre + 0.01 < err_mov_deud:
                    naturaleza = "ACREEDORA"
                    confianza = "ALTA"
                else:
                    naturaleza = "INDETERMINADA"
                    confianza = "BAJA"
            else:
                naturaleza = "INDETERMINADA"
                confianza = "BAJA"
        else:
            # Ninguna ecuación cuadra. Elegimos solo si una es claramente mejor;
            # de cualquier forma la confianza será baja y quedará hallazgo.
            if ed < ea * 0.25:
                naturaleza = "DEUDORA"
                confianza = "BAJA"
            elif ea < ed * 0.25:
                naturaleza = "ACREEDORA"
                confianza = "BAJA"
            else:
                naturaleza = "INDETERMINADA"
                confianza = "BAJA"

        naturalezas.append(naturaleza)
        confianzas.append(confianza)

    r["naturaleza"] = naturalezas
    r["naturaleza_confianza"] = confianzas
    return r


def aplicar_naturaleza_a_movimientos(movs, resumen_naturaleza):
    m = movs.copy()
    # Identificador estable dentro de la ejecución. Permite regresar desde los
    # grupos de conciliación hasta la fila exacta del auxiliar de origen.
    m["movimiento_id"] = np.arange(1, len(m) + 1)
    mapa_nat = resumen_naturaleza.set_index("cuenta_uid")["naturaleza"]
    m["naturaleza"] = m["cuenta_uid"].map(mapa_nat)

    m["efecto_natural"] = np.select(
        [
            m["naturaleza"].eq("DEUDORA"),
            m["naturaleza"].eq("ACREEDORA"),
        ],
        [
            m["cargos"] - m["abonos"],
            m["abonos"] - m["cargos"],
        ],
        default=np.nan,
    )

    m["importe_abs"] = m["efecto_natural"].abs().round(2)
    return m


def marcar_duplicados_exactos(movs):
    m = movs.copy()
    subset = [
        "cuenta_uid", "fecha", "tipo_poliza", "poliza", "concepto_norm",
        "referencia_norm", "cargos", "abonos"
    ]
    m["posible_duplicado_exacto"] = m.duplicated(
        subset=subset, keep=False
    )
    return m


def analizar_saldos(movs, resumen_naturaleza):
    """
    Conciliación por naturaleza:

      saldo_final =
          saldo_inicial
        + efecto natural de movimientos CON referencia
        + efecto natural de movimientos SIN referencia
        + descuadre_origen

    Los hallazgos son independientes; el estado es solo una prioridad visual.
    """
    m = movs.copy()
    r = resumen_naturaleza.copy()

    con_ref = (
        m[m["tiene_referencia"]]
        .groupby("cuenta_uid")["efecto_natural"]
        .sum(min_count=1)
    )
    sin_ref = (
        m[~m["tiene_referencia"]]
        .groupby("cuenta_uid")["efecto_natural"]
        .sum(min_count=1)
    )

    sin_ref_stats = (
        m[~m["tiene_referencia"]]
        .groupby("cuenta_uid")
        .agg(
            n_sin_referencia=("cuenta_uid", "size"),
            cargos_sin_referencia=("cargos", "sum"),
            abonos_sin_referencia=("abonos", "sum"),
        )
    )

    rec_stats = (
        m[m["referencia_recuperada"]]
        .groupby("cuenta_uid")
        .size()
        .rename("n_refs_recuperadas")
    )

    otras_ref_stats = (
        m[m["referencia_tipo"].eq("OTRA_REFERENCIA")]
        .groupby("cuenta_uid")
        .size()
        .rename("n_referencias_libres")
    )

    neg_stats = (
        m[(m["cargos"] < 0) | (m["abonos"] < 0)]
        .groupby("cuenta_uid")
        .size()
        .rename("n_montos_negativos")
    )

    dup_stats = (
        m[m["posible_duplicado_exacto"]]
        .groupby("cuenta_uid")
        .size()
        .rename("n_filas_posible_duplicado")
    )

    r["movs_con_referencia"] = r["cuenta_uid"].map(con_ref).fillna(0.0)
    r["movs_sin_referencia"] = r["cuenta_uid"].map(sin_ref).fillna(0.0)

    r = r.merge(
        sin_ref_stats,
        left_on="cuenta_uid",
        right_index=True,
        how="left",
    )
    for c in [
        "n_sin_referencia", "cargos_sin_referencia",
        "abonos_sin_referencia"
    ]:
        r[c] = r[c].fillna(0)

    r["importe_bruto_sin_referencia"] = (
        r["cargos_sin_referencia"].abs()
        + r["abonos_sin_referencia"].abs()
    )

    r["n_refs_recuperadas"] = r["cuenta_uid"].map(rec_stats).fillna(0).astype(int)
    r["n_referencias_libres"] = (
        r["cuenta_uid"].map(otras_ref_stats).fillna(0).astype(int)
    )
    r["n_montos_negativos"] = (
        r["cuenta_uid"].map(neg_stats).fillna(0).astype(int)
    )
    r["n_filas_posible_duplicado"] = (
        r["cuenta_uid"].map(dup_stats).fillna(0).astype(int)
    )

    r["saldo_esperado_motor"] = (
        r["saldo_inicial"]
        + r["movs_con_referencia"]
        + r["movs_sin_referencia"]
    )
    r["descuadre_origen"] = (
        r["saldo_final_aux"] - r["saldo_esperado_motor"]
    )

    r["cuadra"] = (
        r["descuadre_origen"].abs() <= UMBRAL_TOLERANCIA
    )
    r["tiene_arrastre"] = (
        r["saldo_inicial"].abs() > UMBRAL_TOLERANCIA
    )
    r["tiene_sin_referencia"] = r["n_sin_referencia"] > 0
    r["tiene_montos_negativos"] = r["n_montos_negativos"] > 0

    def estado(row):
        if row["naturaleza"] == "INDETERMINADA":
            return "⚫ Naturaleza indeterminada"
        if abs(row["descuadre_origen"]) > UMBRAL_TOLERANCIA:
            return "🟠 Total fuente ≠ Detalle"
        if row["n_sin_referencia"] > 0:
            return "🔴 Movimientos sin referencia"
        if row["n_montos_negativos"] > 0:
            return "🟣 Montos negativos / reversos"
        return "🟢 OK"

    r["estado"] = r.apply(estado, axis=1)
    return r


# ==============================================================================
# 4. FOLIOS, REFERENCIAS Y CRUCES ARPON
# ==============================================================================

def analizar_folios(movs, fecha_corte):
    """
    Analiza solo referencias que parecen folio documental.
    La antigüedad es desde la PRIMERA fecha observada, NO fecha de vencimiento.
    """
    mv = movs[
        movs["es_folio"]
        & movs["efecto_natural"].notna()
    ].copy()

    if mv.empty:
        return pd.DataFrame(
            columns=[
                "sistema_origen", "empresa", "archivo", "meta_codigo",
                "meta_nombre", "naturaleza", "referencia_norm",
                "primera_fecha", "ultima_fecha", "n_movs",
                "cargos", "abonos", "saldo_natural", "dias",
                "antiguedad_observada", "tipo_saldo",
                "multiples_movimientos", "posible_duplicado_exacto"
            ]
        )

    g = (
        mv.groupby(
            [
                "sistema_origen", "empresa_uid", "empresa", "archivo",
                "cuenta_uid", "cuenta_logica_uid", "meta_codigo", "meta_nombre",
                "naturaleza", "referencia_norm"
            ],
            as_index=False,
        )
        .agg(
            primera_fecha=("fecha", "min"),
            ultima_fecha=("fecha", "max"),
            n_movs=("efecto_natural", "size"),
            cargos=("cargos", "sum"),
            abonos=("abonos", "sum"),
            saldo_natural=("efecto_natural", "sum"),
            posible_duplicado_exacto=(
                "posible_duplicado_exacto", "max"
            ),
        )
    )

    vivos = g[g["saldo_natural"].abs() > UMBRAL_FOLIO].copy()

    corte = pd.Timestamp(fecha_corte)
    vivos["dias"] = (corte - vivos["primera_fecha"]).dt.days

    def bucket(d):
        if pd.isna(d):
            return "sin fecha"
        if d < 0:
            return "fecha posterior al corte"
        if d <= 30:
            return "0-30"
        if d <= 60:
            return "31-60"
        if d <= 90:
            return "61-90"
        return "90+"

    vivos["antiguedad_observada"] = vivos["dias"].apply(bucket)

    def tipo_saldo(row):
        s = row["saldo_natural"]
        nat = row["naturaleza"]
        if nat == "DEUDORA":
            if s > 0:
                return "🔵 Saldo deudor pendiente"
            return "🔴 Saldo contrario a naturaleza (acreedor)"
        if nat == "ACREEDORA":
            if s > 0:
                return "🟣 Saldo acreedor pendiente / por aplicar"
            return "🔴 Saldo contrario a naturaleza (deudor)"
        return "⚫ Naturaleza indeterminada"

    vivos["tipo_saldo"] = vivos.apply(tipo_saldo, axis=1)
    vivos["multiples_movimientos"] = vivos["n_movs"] > 2

    return vivos.sort_values(
        ["dias", "saldo_natural"], ascending=[False, False]
    )


def detectar_cruces_por_referencia(movs):
    """
    Busca el mismo folio en cuentas distintas, pero SOLO dentro del mismo
    empresa ARPON, evitando cruces entre empresas distintas.
    """
    mv = movs[
        movs["es_folio"]
        & movs["efecto_natural"].notna()
    ].copy()

    if mv.empty:
        return pd.DataFrame()

    por_cuenta = (
        mv.groupby(
            [
                "sistema_origen", "empresa_uid", "empresa",
                "referencia_norm", "cuenta_logica_uid",
                "meta_codigo", "meta_nombre", "naturaleza"
            ],
            as_index=False,
        )
        .agg(
            cargos=("cargos", "sum"),
            abonos=("abonos", "sum"),
            efecto_natural=("efecto_natural", "sum"),
            n_movs=("efecto_natural", "size"),
            archivos=("archivo", lambda x: " | ".join(sorted(set(map(str, x))))),
        )
    )

    claves_ref = ["sistema_origen", "empresa_uid", "referencia_norm"]
    nivel_ref = (
        por_cuenta.groupby(claves_ref)
        .agg(
            num_cuentas=("cuenta_logica_uid", "nunique"),
            hay_positivo=("efecto_natural", lambda x: (x > UMBRAL_FOLIO).any()),
            hay_negativo=("efecto_natural", lambda x: (x < -UMBRAL_FOLIO).any()),
            neto_global=("efecto_natural", "sum"),
        )
        .reset_index()
    )

    refs = nivel_ref[
        (nivel_ref["num_cuentas"] > 1)
        & nivel_ref["hay_positivo"]
        & nivel_ref["hay_negativo"]
    ].copy()

    if refs.empty:
        return pd.DataFrame()

    detalle = por_cuenta.merge(
        refs[claves_ref + ["num_cuentas", "neto_global"]],
        on=claves_ref,
        how="inner",
    )
    detalle["amarre_aprox"] = detalle["neto_global"].abs() <= UMBRAL_TOLERANCIA
    detalle["nivel_evidencia"] = np.where(
        detalle["amarre_aprox"], "ALTA - neto aproximado a cero",
        "MEDIA - efectos opuestos con remanente"
    )
    return detalle.sort_values(
        ["sistema_origen", "empresa_uid", "referencia_norm", "efecto_natural"],
        ascending=[True, True, True, False],
    )


def detectar_coincidencias_por_evidencia(movs):
    """
    Coincidencias por fecha + concepto + importe dentro del mismo sistema y
    empresa. Se clasifican por calidad del amarre para evitar presentar como
    conciliación fuerte un grupo muchos-a-muchos con remanente.
    """
    mv = movs[
        movs["efecto_natural"].notna()
        & (movs["importe_abs"] > UMBRAL_FOLIO)
        & movs["concepto_norm"].ne("")
    ].copy()

    if mv.empty:
        return pd.DataFrame()

    mv = mv[mv["efecto_natural"].abs() > UMBRAL_FOLIO].copy()
    claves = [
        "sistema_origen", "empresa_uid", "fecha", "concepto_norm", "importe_abs"
    ]

    grupos = (
        mv.groupby(claves)
        .agg(
            num_cuentas=("cuenta_logica_uid", "nunique"),
            hay_positivo=("efecto_natural", lambda x: (x > 0).any()),
            hay_negativo=("efecto_natural", lambda x: (x < 0).any()),
            n_movs_grupo=("efecto_natural", "size"),
            n_positivos=("efecto_natural", lambda x: int((x > 0).sum())),
            n_negativos=("efecto_natural", lambda x: int((x < 0).sum())),
            neto_grupo=("efecto_natural", "sum"),
        )
        .reset_index()
    )

    validos = grupos[
        (grupos["num_cuentas"] > 1)
        & grupos["hay_positivo"]
        & grupos["hay_negativo"]
    ].copy()

    if validos.empty:
        return pd.DataFrame()

    validos["amarre_aprox"] = validos["neto_grupo"].abs() <= UMBRAL_TOLERANCIA
    validos["nivel_evidencia"] = np.select(
        [
            validos["amarre_aprox"]
            & validos["n_positivos"].eq(1)
            & validos["n_negativos"].eq(1),
            validos["amarre_aprox"],
        ],
        [
            "ALTA - correspondencia 1:1",
            "MEDIA - neto cero con múltiples movimientos",
        ],
        default="BAJA - coincidencia parcial con remanente",
    )
    validos["evidencia_id"] = np.arange(1, len(validos) + 1)

    det = mv.merge(validos, on=claves, how="inner")
    cols = [
        "evidencia_id", "nivel_evidencia", "amarre_aprox",
        "movimiento_id", "fila_origen", "sistema_origen", "empresa_uid",
        "empresa", "fecha", "concepto", "concepto_norm",
        "importe_abs", "archivo", "cuenta_uid", "cuenta_logica_uid",
        "meta_codigo", "meta_nombre", "naturaleza",
        "referencia_original", "referencia_norm", "referencia_fuente",
        "cargos", "abonos", "efecto_natural",
        "num_cuentas", "n_movs_grupo", "n_positivos", "n_negativos", "neto_grupo"
    ]
    return det[cols].sort_values(
        ["evidencia_id", "efecto_natural"], ascending=[True, False]
    )


def marcar_movimientos_conciliacion(movs, cruces_ref, evidencias):
    """
    Regresa los movimientos con una marca auditable de conciliación.

    CONCILIADO:
      - el folio tiene efectos opuestos y neto aproximado a cero; o
      - la evidencia por fecha + concepto + importe tiene neto cero.

    REVISAR:
      - existen efectos opuestos, pero el grupo conserva un remanente.

    No elimina ni compensa movimientos; únicamente agrega trazabilidad para
    poder marcar la fila original del auxiliar ARPON.
    """
    m = movs.copy()
    if "movimiento_id" not in m.columns:
        m["movimiento_id"] = np.arange(1, len(m) + 1)

    registros = {
        int(mid): {
            "estado": "SIN MARCA",
            "nivel": "",
            "criterios": [],
            "codigos": [],
        }
        for mid in m["movimiento_id"]
    }
    rango_estado = {"SIN MARCA": 0, "REVISAR": 1, "CONCILIADO": 2}
    rango_nivel = {"": 0, "BAJA": 1, "MEDIA": 2, "ALTA": 3}

    def registrar(ids, estado, nivel, criterio, codigo):
        nivel_base = str(nivel).split(" - ", 1)[0].strip().upper()
        if nivel_base not in rango_nivel:
            nivel_base = "BAJA"

        for mid in ids:
            reg = registros.get(int(mid))
            if reg is None:
                continue
            if rango_estado[estado] > rango_estado[reg["estado"]]:
                reg["estado"] = estado
            if rango_nivel[nivel_base] > rango_nivel[reg["nivel"]]:
                reg["nivel"] = nivel_base
            if criterio not in reg["criterios"]:
                reg["criterios"].append(criterio)
            if codigo not in reg["codigos"]:
                reg["codigos"].append(codigo)

    if cruces_ref is not None and not cruces_ref.empty:
        claves = [
            "sistema_origen", "empresa_uid", "referencia_norm",
            "amarre_aprox", "nivel_evidencia",
        ]
        for _, grupo in cruces_ref[claves].drop_duplicates().iterrows():
            mask = (
                m["sistema_origen"].eq(grupo["sistema_origen"])
                & m["empresa_uid"].eq(grupo["empresa_uid"])
                & m["referencia_norm"].eq(grupo["referencia_norm"])
                & m["es_folio"]
                & (m["efecto_natural"].abs() > UMBRAL_FOLIO)
            )
            registrar(
                m.loc[mask, "movimiento_id"],
                "CONCILIADO" if bool(grupo["amarre_aprox"]) else "REVISAR",
                grupo["nivel_evidencia"],
                "FOLIO",
                f"REF:{grupo['referencia_norm']}",
            )

    if evidencias is not None and not evidencias.empty:
        grupos_evidencia = evidencias[
            ["evidencia_id", "nivel_evidencia", "amarre_aprox"]
        ].drop_duplicates()
        for _, grupo in grupos_evidencia.iterrows():
            ids = evidencias.loc[
                evidencias["evidencia_id"].eq(grupo["evidencia_id"]),
                "movimiento_id",
            ].drop_duplicates()
            nivel = str(grupo["nivel_evidencia"]).split(" - ", 1)[0].upper()
            if nivel == "ALTA":
                criterio = "EVIDENCIA 1:1"
            elif nivel == "MEDIA":
                criterio = "EVIDENCIA GRUPAL"
            else:
                criterio = "COINCIDENCIA PARCIAL"
            registrar(
                ids,
                "CONCILIADO" if bool(grupo["amarre_aprox"]) else "REVISAR",
                grupo["nivel_evidencia"],
                criterio,
                f"EVD:{int(grupo['evidencia_id'])}",
            )

    m["conciliacion_estado"] = m["movimiento_id"].map(
        lambda mid: registros[int(mid)]["estado"]
    )
    m["conciliacion_nivel"] = m["movimiento_id"].map(
        lambda mid: registros[int(mid)]["nivel"]
    )
    m["conciliacion_criterio"] = m["movimiento_id"].map(
        lambda mid: " + ".join(registros[int(mid)]["criterios"])
    )
    m["conciliacion_codigo"] = m["movimiento_id"].map(
        lambda mid: " | ".join(registros[int(mid)]["codigos"])
    )
    m["conciliacion_marcada"] = m["conciliacion_estado"].ne("SIN MARCA")
    return m


def _buscar_fila_encabezado_arpon(ws):
    limite = min(ws.max_row, 60)
    for fila in range(1, limite + 1):
        vals = [texto_norm(ws.cell(fila, col).value) for col in range(1, 8)]
        if (
            vals[0] == "POLIZA"
            and vals[1] == "FECHA"
            and vals[2].startswith("DOCTO")
            and vals[3] == "CONCEPTO"
            and vals[4] == "CARGO"
            and vals[5] == "ABONO"
            and vals[6] == "SALDO"
        ):
            return fila
    raise ValueError("No se encontró el encabezado ARPON en el libro de origen.")


def _libro_desde_archivo(file_bytes, file_name):
    lower = file_name.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return load_workbook(
            BytesIO(file_bytes),
            keep_vba=lower.endswith(".xlsm"),
            keep_links=True,
        )

    # CSV/XLS se convierten a XLSX para poder entregar el marcado visual.
    raw = cargar_archivo_robusto(file_bytes, file_name)
    wb = Workbook()
    ws = wb.active
    ws.title = "Auxiliar ARPON"
    for fila_idx, valores in enumerate(
        raw.itertuples(index=False, name=None), start=1
    ):
        for col_idx, valor in enumerate(valores, start=1):
            if pd.isna(valor):
                valor = None
            elif isinstance(valor, pd.Timestamp):
                valor = valor.to_pydatetime()
            ws.cell(fila_idx, col_idx, valor)
    return wb


def construir_auxiliar_marcado(file_bytes, file_name, marcas):
    """Conserva el auxiliar y agrega color + código en la fila conciliada."""
    wb = _libro_desde_archivo(file_bytes, file_name)
    ws = wb.worksheets[0]
    fila_header = _buscar_fila_encabezado_arpon(ws)

    columna_estado = None
    for col in range(8, ws.max_column + 1):
        if texto_norm(ws.cell(fila_header, col).value) == "CONCILIACION":
            columna_estado = col
            break
    if columna_estado is None:
        columna_estado = max(8, ws.max_column + 1)

    celda_header = ws.cell(fila_header, columna_estado)
    fuente_header = ws.cell(fila_header, 7)
    if celda_header.coordinate != fuente_header.coordinate:
        celda_header._style = copy(fuente_header._style)
        celda_header.number_format = fuente_header.number_format
        celda_header.alignment = copy(fuente_header.alignment)
    celda_header.value = "Conciliación"
    celda_header.font = copy(celda_header.font)
    celda_header.font = Font(
        name=celda_header.font.name,
        size=celda_header.font.size,
        bold=True,
        color=celda_header.font.color,
    )
    celda_header.alignment = Alignment(horizontal="center", vertical="center")

    verde = PatternFill("solid", fgColor="C6EFCE")
    verde_suave = PatternFill("solid", fgColor="EAF4E3")
    amarillo = PatternFill("solid", fgColor="FFEB9C")
    amarillo_suave = PatternFill("solid", fgColor="FFF7D6")

    leyenda = [
        (1, "Marca de conciliación"),
        (2, "✓ Verde = conciliado"),
        (3, "⚠ Amarillo = revisar remanente"),
    ]
    for fila, texto in leyenda:
        celda = ws.cell(fila, columna_estado)
        if es_vacio(celda.value):
            celda.value = texto
            celda.font = Font(bold=(fila == 1), color="1F1F1F", size=10)
            celda.fill = verde if fila == 2 else amarillo if fila == 3 else verde_suave

    for _, marca in marcas.sort_values("fila_origen").iterrows():
        fila = int(marca["fila_origen"])
        if fila < 1 or fila > ws.max_row:
            continue

        conciliado = marca["conciliacion_estado"] == "CONCILIADO"
        simbolo = "✓" if conciliado else "⚠"
        texto = (
            f"{simbolo} {marca['conciliacion_estado']} · "
            f"{marca['conciliacion_criterio']} · {marca['conciliacion_codigo']}"
        )
        celda = ws.cell(fila, columna_estado)
        celda.value = texto
        celda.fill = verde if conciliado else amarillo
        celda.font = Font(
            bold=True,
            color="006100" if conciliado else "9C6500",
            size=10,
        )
        celda.alignment = Alignment(vertical="center", wrap_text=False)

        relleno_fila = verde_suave if conciliado else amarillo_suave
        for col in range(1, columna_estado):
            origen = ws.cell(fila, col)
            if origen.fill is None or origen.fill.fill_type is None:
                origen.fill = relleno_fila

    letra_estado = ws.cell(1, columna_estado).column_letter
    ws.column_dimensions[letra_estado].width = max(
        ws.column_dimensions[letra_estado].width or 0,
        58,
    )

    output = BytesIO()
    extension = ".xlsm" if file_name.lower().endswith(".xlsm") else ".xlsx"
    wb.save(output)
    nombre_salida = f"{Path(file_name).stem}_MARCADO{extension}"
    return output.getvalue(), nombre_salida


@st.cache_data(show_spinner=False)
def construir_descarga_auxiliares_marcados(archivos, movs):
    """Devuelve un XLSX/XLSM si es uno, o un ZIP si se cargaron varios."""
    resultados = []
    nombres_usados = set()

    for file_name, file_bytes in archivos:
        marcas = movs[
            movs["archivo"].eq(file_name)
            & movs["conciliacion_marcada"]
        ].copy()
        data, nombre = construir_auxiliar_marcado(file_bytes, file_name, marcas)

        base = Path(nombre).stem
        extension = Path(nombre).suffix
        candidato = nombre
        i = 2
        while candidato in nombres_usados:
            candidato = f"{base}_{i}{extension}"
            i += 1
        nombres_usados.add(candidato)
        resultados.append((candidato, data))

    if len(resultados) == 1:
        nombre, data = resultados[0]
        mime = (
            "application/vnd.ms-excel.sheet.macroEnabled.12"
            if nombre.lower().endswith(".xlsm")
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return data, nombre, mime

    output = BytesIO()
    with ZipFile(output, mode="w", compression=ZIP_DEFLATED) as zf:
        for nombre, data in resultados:
            zf.writestr(nombre, data)
    return output.getvalue(), "auxiliares_ARPON_MARCADOS.zip", "application/zip"


def tabla_referencias(movs):
    cols = [
        "sistema_origen", "empresa", "archivo", "fila_origen", "fecha",
        "meta_codigo", "meta_nombre",
        "concepto", "referencia_original", "referencia_norm",
        "referencia_tipo", "referencia_fuente", "referencia_recuperada",
        "cargos", "abonos", "naturaleza", "efecto_natural",
        "conciliacion_estado", "conciliacion_nivel",
        "conciliacion_criterio", "conciliacion_codigo",
    ]
    return movs[cols].copy()


# ==============================================================================
# 5. UI
# ==============================================================================

def main():
    st.set_page_config(
        page_title="Auditoría ARPON · Hotel Quartz",
        layout="wide",
        page_icon="🛡️",
    )

    st.title("🛡️ Auditoría de Saldos ARPON · Hotel Quartz")
    st.caption(f"Motor v{APP_VERSION}")

    st.markdown(
        """
        Motor exclusivo para auxiliares **ARPON de Hotel Quartz**.

        - valida la estructura **Póliza | Fecha | Docto. | Concepto | Cargo | Abono | Saldo**;
        - identifica empresa, periodo, cuenta y saldo inicial desde el propio reporte;
        - valida **cargos, abonos y saldo acumulado movimiento por movimiento**;
        - detecta automáticamente la **naturaleza deudora o acreedora**;
        - conserva y normaliza documentos/folios sin destruir prefijos;
        - puede recuperar un folio desde **Concepto** cuando Docto. está vacío;
        - identifica movimientos sin referencia, reversos y posibles duplicados;
        - los cruces se realizan únicamente dentro de la **misma empresa ARPON**;
        - genera una copia del auxiliar con **color y código de conciliación**.
        """
    )

    uploaded_files = st.file_uploader(
        "📂 Sube uno o varios Auxiliares de Cuentas de ARPON (Excel o CSV)",
        type=["xlsx", "xls", "xlsm", "csv"],
        accept_multiple_files=True,
    )

    if not uploaded_files:
        st.info("Esperando archivo(s)...")
        return

    movs_lista = []
    resumen_lista = []
    diags = []
    errores = []

    with st.spinner("Procesando y validando auxiliares..."):
        for uf in uploaded_files:
            try:
                movs_i, resumen_i, diag_i = procesar_archivo_engine(
                    uf.getvalue(), uf.name
                )
                movs_lista.append(movs_i)
                resumen_lista.append(resumen_i)
                diags.append(diag_i)
            except Exception as e:
                errores.append(f"**{uf.name}:** {e}")

    if errores:
        st.error(
            "⚠️ **No se certifica la lectura. Corrige o revisa estos archivos antes "
            "de usar resultados:**\n\n"
            + "\n\n".join(f"- {x}" for x in errores)
        )
        st.stop()

    movs = pd.concat(movs_lista, ignore_index=True)
    resumen = pd.concat(resumen_lista, ignore_index=True)

    # Evitar que cargar dos veces la misma cuenta pase desapercibido.
    repetidas = (
        resumen.groupby(["empresa_uid", "meta_codigo"])["archivo"]
        .nunique()
        .loc[lambda s: s > 1]
    )
    if not repetidas.empty:
        st.warning(
            "⚠️ Hay códigos de cuenta presentes en más de un archivo. "
            "No necesariamente es un error, pero revisa que no hayas cargado "
            "dos periodos o copias de la misma cuenta: "
            + ", ".join(str(x) for x in repetidas.index)
        )

    resumen_nat = detectar_naturaleza(resumen, movs)
    movs = aplicar_naturaleza_a_movimientos(movs, resumen_nat)
    movs = marcar_duplicados_exactos(movs)
    df_cruces_ref = detectar_cruces_por_referencia(movs)
    df_evidencia = detectar_coincidencias_por_evidencia(movs)
    movs = marcar_movimientos_conciliacion(
        movs, df_cruces_ref, df_evidencia
    )
    df_audit = analizar_saldos(movs, resumen_nat)

    # --------------------------------------------------------------------------
    # Validación visible de lectura
    # --------------------------------------------------------------------------
    st.divider()
    st.subheader("✅ Validación de lectura")

    diag_df = pd.DataFrame(diags)

    st.caption("Sistema contable: ARPON")
    empresas_arpon = sorted(
        diag_df["empresa"].dropna().astype(str)
        .loc[lambda x: x.str.strip().ne("")].unique()
    )
    if empresas_arpon:
        st.caption("Empresa detectada: " + " | ".join(empresas_arpon))

    n_archivos = len(diag_df)
    n_cuentas = len(df_audit)
    n_movs = len(movs)

    amarres_false = diag_df["amarre_gran_total"].eq(False).sum()
    if amarres_false:
        st.warning(
            f"{amarres_false} archivo(s) no amarran la suma de saldos por cuenta "
            "contra el saldo final reportado por ARPON. Revisa si el reporte contiene "
            "agrupaciones adicionales."
        )
    else:
        st.success(
            f"Lectura estructural validada: **{n_archivos} archivo(s)** · "
            f"**{n_cuentas} cuenta(s)** · **{n_movs:,} movimientos**. "
            "La estructura, los totales disponibles y las secuencias de saldo fueron validados."
        )

    for _, d in diag_df.iterrows():
        gt_txt = (
            f"${d['gran_total']:,.2f}"
            if pd.notna(d.get("gran_total"))
            else "no detectado"
        )
        amarre = d.get("amarre_gran_total")
        if pd.isna(amarre):
            estado_amarre = "ℹ️"
        elif bool(amarre):
            estado_amarre = "✅"
        else:
            estado_amarre = "⚠️"
        st.caption(
            f"{estado_amarre} {d['archivo']} · {'ARPON'} [{d.get('formato', 'N/D')}]: "
            f"{int(d['n_headers'])} cuenta(s), {int(d['n_movs']):,} movimientos, "
            f"Total reportado {gt_txt}."
        )

    # --------------------------------------------------------------------------
    # KPIs
    # --------------------------------------------------------------------------
    gran_totales_validos = pd.to_numeric(
        diag_df.get("gran_total", pd.Series(dtype=float)), errors="coerce"
    ).dropna()
    saldo_total = (
        float(gran_totales_validos.sum())
        if len(gran_totales_validos) == len(diag_df) and len(diag_df) > 0
        else float(df_audit["saldo_final_aux"].sum())
    )
    bruto_sin_ref = df_audit["importe_bruto_sin_referencia"].sum()
    descuadre_abs = df_audit["descuadre_origen"].abs().sum()
    n_sin_ref = int(df_audit["n_sin_referencia"].sum())
    n_revisar = int((df_audit["estado"] != "🟢 OK").sum())

    n_refs_cruce = (
        int(df_cruces_ref["referencia_norm"].nunique())
        if not df_cruces_ref.empty else 0
    )
    n_evidencias = (
        int(df_evidencia["evidencia_id"].nunique())
        if not df_evidencia.empty else 0
    )
    n_partidas_conciliadas = int(
        movs["conciliacion_estado"].eq("CONCILIADO").sum()
    )
    n_partidas_revisar = int(
        movs["conciliacion_estado"].eq("REVISAR").sum()
    )

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Saldo total reportado", f"${saldo_total:,.2f}")
    k2.metric(
        "Movs sin referencia",
        f"{n_sin_ref:,}",
        help=f"Importe bruto involucrado: ${bruto_sin_ref:,.2f}",
    )
    k3.metric(
        "Descuadre absoluto",
        f"${descuadre_abs:,.2f}",
        help="Suma de valores absolutos por cuenta; evita compensar + y -.",
    )
    k4.metric("Cruces por folio", n_refs_cruce)
    k5.metric("Cruces por evidencia", n_evidencias)
    k6.metric("Cuentas a revisar", n_revisar)
    st.caption(
        f"Marcas para auxiliares: {n_partidas_conciliadas:,} partida(s) "
        f"conciliada(s) y {n_partidas_revisar:,} partida(s) con remanente "
        "para revisión."
    )

    # Fecha de corte
    fmax = movs["fecha"].max()
    corte_default = (
        fmax.date() if pd.notna(fmax) else pd.Timestamp.now().date()
    )
    corte = st.date_input(
        "📅 Fecha de corte para antigüedad observada",
        value=corte_default,
        help=(
            "La antigüedad se calcula desde la primera fecha observada del folio. "
            "No equivale a días vencidos si no existe fecha de vencimiento."
        ),
    )
    folios = analizar_folios(movs, corte)

    # --------------------------------------------------------------------------
    # Pestañas
    # --------------------------------------------------------------------------
    tabs = st.tabs(
        [
            "🔎 Hallazgos",
            "🚦 Semáforo",
            "📑 Folios",
            "✅ Conciliación marcada",
            "🏷️ Referencias",
            "📉 Gráficos",
            "🧪 Diagnóstico",
        ]
    )

    # --------------------------------------------------------------------------
    # Hallazgos
    # --------------------------------------------------------------------------
    with tabs[0]:
        st.subheader("🔎 Hallazgos priorizados")
        st.caption(
            "Los hallazgos son independientes. Una cuenta puede tener más de uno."
        )

        sin_ref_movs = movs[~movs["tiene_referencia"]].copy()
        refs_rec = movs[movs["referencia_recuperada"]].copy()
        negativos = movs[(movs["cargos"] < 0) | (movs["abonos"] < 0)].copy()
        duplicados = movs[movs["posible_duplicado_exacto"]].copy()
        descuadres = df_audit[
            df_audit["descuadre_origen"].abs() > UMBRAL_TOLERANCIA
        ].copy()
        contrarios = folios[
            folios["tipo_saldo"].str.contains("contrario", case=False, na=False)
        ].copy()
        viejos = folios[
            folios["antiguedad_observada"].eq("90+")
        ].copy()

        h1, h2, h3, h4, h5, h6 = st.columns(6)
        h1.metric("Cuentas descuadre", len(descuadres))
        h2.metric("Movs sin ref", len(sin_ref_movs))
        h3.metric("Refs recuperadas", len(refs_rec))
        h4.metric("Montos negativos", len(negativos))
        h5.metric("Posibles duplicados", len(duplicados))
        h6.metric("Folios 90+ observados", len(viejos))

        if len(descuadres):
            st.markdown("#### 🟠 Descuadre contra el saldo final reportado por ARPON")
            st.dataframe(
                descuadres[
                    [
                        "sistema_origen", "empresa", "archivo", "meta_codigo",
                        "meta_nombre", "naturaleza", "saldo_final_aux", "saldo_esperado_motor",
                        "descuadre_origen"
                    ]
                ],
                use_container_width=True,
                hide_index=True,
            )

        if len(sin_ref_movs):
            st.markdown("#### 🔴 Movimientos realmente sin referencia")
            st.caption(
                "No había referencia en la columna y tampoco fue posible recuperar "
                "un folio documental inequívoco desde Concepto."
            )
            st.dataframe(
                sin_ref_movs[
                    [
                        "archivo", "fila_origen", "fecha", "meta_codigo",
                        "concepto", "cargos", "abonos", "efecto_natural"
                    ]
                ],
                use_container_width=True,
                hide_index=True,
            )

        if len(refs_rec):
            st.markdown("#### 🟡 Referencias recuperadas desde Concepto")
            st.caption(
                "No se consideran 'sin referencia', pero se muestran para trazabilidad."
            )
            st.dataframe(
                refs_rec[
                    [
                        "archivo", "fila_origen", "fecha", "meta_codigo",
                        "concepto", "referencia_original", "referencia_norm",
                        "cargos", "abonos"
                    ]
                ],
                use_container_width=True,
                hide_index=True,
            )

        if len(negativos):
            st.markdown("#### 🟣 Montos negativos / reversos")
            st.caption(
                "Se señalan como movimiento especial; no se reinterpretan como abono."
            )
            st.dataframe(
                negativos[
                    [
                        "archivo", "fila_origen", "fecha", "meta_codigo",
                        "concepto", "referencia_norm", "cargos", "abonos",
                        "efecto_natural"
                    ]
                ],
                use_container_width=True,
                hide_index=True,
            )

        if len(duplicados):
            st.markdown("#### 🔁 Posibles duplicados exactos")
            st.caption(
                "Misma cuenta, fecha, tipo, póliza, concepto, referencia y monto. "
                "Es un indicador para revisión, no una conclusión automática."
            )
            st.dataframe(
                duplicados[
                    [
                        "archivo", "fila_origen", "fecha", "meta_codigo",
                        "tipo_poliza", "poliza", "concepto",
                        "referencia_norm", "cargos", "abonos"
                    ]
                ],
                use_container_width=True,
                hide_index=True,
            )

        if len(contrarios):
            st.markdown("#### ⚠️ Folios con saldo contrario a la naturaleza")
            st.dataframe(
                contrarios[
                    [
                        "archivo", "meta_codigo", "meta_nombre", "naturaleza",
                        "referencia_norm", "primera_fecha", "dias",
                        "cargos", "abonos", "saldo_natural", "tipo_saldo"
                    ]
                ],
                use_container_width=True,
                hide_index=True,
            )

        if not any(
            [
                len(descuadres), len(sin_ref_movs), len(refs_rec),
                len(negativos), len(duplicados), len(contrarios)
            ]
        ):
            st.success("Sin hallazgos relevantes con los criterios actuales.")

    # --------------------------------------------------------------------------
    # Semáforo
    # --------------------------------------------------------------------------
    with tabs[1]:
        st.subheader("🚦 Conciliación por cuenta")
        solo_problemas = st.toggle(
            "Ver solo cuentas con hallazgos",
            value=False,
            key="solo_problemas",
        )
        show = (
            df_audit[df_audit["estado"] != "🟢 OK"]
            if solo_problemas else df_audit
        )

        cols = [
            "sistema_origen", "empresa", "archivo", "meta_codigo",
            "meta_nombre", "naturaleza", "naturaleza_confianza", "estado", "saldo_inicial",
            "total_cargos", "total_abonos", "saldo_final_aux",
            "movs_con_referencia", "movs_sin_referencia",
            "n_sin_referencia", "importe_bruto_sin_referencia",
            "n_refs_recuperadas", "n_referencias_libres",
            "n_montos_negativos", "descuadre_origen"
        ]
        st.dataframe(
            show[cols],
            use_container_width=True,
            hide_index=True,
            column_config={
                "saldo_inicial": st.column_config.NumberColumn(
                    "Saldo inicial", format="$%.2f"
                ),
                "total_cargos": st.column_config.NumberColumn(
                    "Cargos", format="$%.2f"
                ),
                "total_abonos": st.column_config.NumberColumn(
                    "Abonos", format="$%.2f"
                ),
                "saldo_final_aux": st.column_config.NumberColumn(
                    "Saldo final ARPON", format="$%.2f"
                ),
                "movs_con_referencia": st.column_config.NumberColumn(
                    "Efecto con referencia", format="$%.2f"
                ),
                "movs_sin_referencia": st.column_config.NumberColumn(
                    "Efecto sin referencia", format="$%.2f"
                ),
                "importe_bruto_sin_referencia": st.column_config.NumberColumn(
                    "Bruto sin referencia", format="$%.2f"
                ),
                "descuadre_origen": st.column_config.NumberColumn(
                    "Descuadre", format="$%.2f"
                ),
            },
        )

    # --------------------------------------------------------------------------
    # Folios
    # --------------------------------------------------------------------------
    with tabs[2]:
        st.subheader("📑 Folios documentales abiertos")
        st.caption(
            "Solo incluye referencias con forma documental reconocible. "
            "La antigüedad es observada desde la primera fecha del folio, "
            "no fecha contractual de vencimiento."
        )

        orden = ["0-30", "31-60", "61-90", "90+"]
        positivos = folios[folios["saldo_natural"] > 0].copy()
        aging = (
            positivos.groupby("antiguedad_observada")["saldo_natural"]
            .agg(num_folios="count", saldo="sum")
            .reindex(orden)
            .fillna(0)
            .reset_index()
        )
        st.dataframe(aging, use_container_width=True, hide_index=True)

        if not folios.empty:
            nat_sel = st.multiselect(
                "Naturaleza",
                sorted(folios["naturaleza"].dropna().unique()),
                default=sorted(folios["naturaleza"].dropna().unique()),
            )
            edades_sel = st.multiselect(
                "Antigüedad observada",
                orden,
                default=orden,
            )
            fv = folios[
                folios["naturaleza"].isin(nat_sel)
                & folios["antiguedad_observada"].isin(edades_sel)
            ]
        else:
            fv = folios

        st.dataframe(
            fv,
            use_container_width=True,
            hide_index=True,
        )

    # --------------------------------------------------------------------------
    # Cruces / conciliación
    # --------------------------------------------------------------------------
    with tabs[3]:
        st.subheader("🔀 Cruces y conciliación entre cuentas")
        c1, c2 = st.columns(2)
        c1.metric("Partidas conciliadas", f"{n_partidas_conciliadas:,}")
        c2.metric("Coincidencias a revisar", f"{n_partidas_revisar:,}")
        st.caption(
            "En pantalla y en el auxiliar descargado, verde significa grupo con neto aproximado "
            "a cero; amarillo significa efectos opuestos con remanente y requiere "
            "revisión."
        )

        st.markdown("#### Partidas marcadas")
        st.caption(
            "Estas son exactamente las filas que recibirán color y código en el "
            "auxiliar descargado."
        )
        partidas_pantalla = movs[movs["conciliacion_marcada"]].copy()
        if partidas_pantalla.empty:
            st.info(
                "No hay partidas para marcar. Carga al mismo tiempo los auxiliares "
                "de las cuentas que deseas conciliar."
            )
        else:
            partidas_pantalla = partidas_pantalla[
                [
                    "conciliacion_estado", "conciliacion_nivel",
                    "conciliacion_criterio", "conciliacion_codigo",
                    "archivo", "fila_origen", "fecha", "meta_codigo",
                    "poliza", "referencia_original", "concepto",
                    "cargos", "abonos", "efecto_natural",
                ]
            ].sort_values(
                ["conciliacion_estado", "archivo", "fila_origen"]
            )
            partidas_pantalla = partidas_pantalla.rename(
                columns={
                    "conciliacion_estado": "Estado",
                    "conciliacion_nivel": "Nivel",
                    "conciliacion_criterio": "Criterio",
                    "conciliacion_codigo": "Código",
                    "archivo": "Archivo",
                    "fila_origen": "Fila ARPON",
                    "fecha": "Fecha",
                    "meta_codigo": "Cuenta",
                    "poliza": "Póliza",
                    "referencia_original": "Docto.",
                    "concepto": "Concepto",
                    "cargos": "Cargo",
                    "abonos": "Abono",
                    "efecto_natural": "Efecto natural",
                }
            )

            def color_partida(row):
                if row["Estado"] == "CONCILIADO":
                    estilo = "background-color: #EAF4E3; color: #006100;"
                else:
                    estilo = "background-color: #FFF7D6; color: #9C6500;"
                return [estilo] * len(row)

            tabla_marcada = (
                partidas_pantalla.style
                .apply(color_partida, axis=1)
                .format(
                    {
                        "Cargo": "${:,.2f}",
                        "Abono": "${:,.2f}",
                        "Efecto natural": "${:,.2f}",
                    },
                    na_rep="",
                )
            )
            st.dataframe(
                tabla_marcada,
                use_container_width=True,
                hide_index=True,
                height=min(620, 85 + 35 * len(partidas_pantalla)),
            )

        st.markdown("#### A. Cruces por el mismo folio")
        if df_cruces_ref.empty:
            st.info(
                "No se encontraron folios idénticos con efectos opuestos "
                "entre cuentas cargadas."
            )
        else:
            st.dataframe(
                df_cruces_ref,
                use_container_width=True,
                hide_index=True,
            )

        st.markdown("#### B. Coincidencias fuertes aunque el folio sea diferente")
        st.caption(
            "Misma fecha + mismo concepto + mismo importe absoluto + "
            "efecto natural opuesto entre cuentas. Es evidencia para conciliar; "
            "no se basa en similitud difusa de nombres."
        )
        if df_evidencia.empty:
            st.info(
                "No se encontraron coincidencias fuertes entre las cuentas cargadas."
            )
        else:
            st.success(
                f"Se encontraron {df_evidencia['evidencia_id'].nunique():,} "
                "grupo(s) de evidencia."
            )
            st.dataframe(
                df_evidencia,
                use_container_width=True,
                hide_index=True,
            )

    # --------------------------------------------------------------------------
    # Referencias
    # --------------------------------------------------------------------------
    with tabs[4]:
        st.subheader("🏷️ Auditoría de referencias")
        refs = tabla_referencias(movs)

        tipos = (
            refs["referencia_tipo"].fillna("VACIA").value_counts()
            .rename_axis("tipo")
            .reset_index(name="movimientos")
        )
        st.dataframe(tipos, use_container_width=True, hide_index=True)

        filtro_tipo = st.multiselect(
            "Tipo de referencia",
            sorted(refs["referencia_tipo"].dropna().unique()),
            default=sorted(refs["referencia_tipo"].dropna().unique()),
        )
        refs_show = refs[refs["referencia_tipo"].isin(filtro_tipo)]
        st.dataframe(
            refs_show,
            use_container_width=True,
            hide_index=True,
        )

    # --------------------------------------------------------------------------
    # Gráficos
    # --------------------------------------------------------------------------
    with tabs[5]:
        st.subheader("📉 Composición del saldo por naturaleza")

        saldo_ini = df_audit["saldo_inicial"].sum()
        con_ref = df_audit["movs_con_referencia"].sum()
        sin_ref = df_audit["movs_sin_referencia"].sum()
        desc = df_audit["descuadre_origen"].sum()

        fig = go.Figure(
            data=[
                go.Bar(
                    name="Saldo inicial",
                    x=["Saldo total"],
                    y=[saldo_ini],
                ),
                go.Bar(
                    name="Efecto con referencia",
                    x=["Saldo total"],
                    y=[con_ref],
                ),
                go.Bar(
                    name="Efecto sin referencia",
                    x=["Saldo total"],
                    y=[sin_ref],
                ),
                go.Bar(
                    name="Descuadre",
                    x=["Saldo total"],
                    y=[desc],
                ),
            ]
        )
        fig.update_layout(
            barmode="relative",
            title="Composición del saldo reportado",
            yaxis_title="Monto",
        )
        st.plotly_chart(fig, use_container_width=True)

    # --------------------------------------------------------------------------
    # Diagnóstico
    # --------------------------------------------------------------------------
    with tabs[6]:
        st.subheader("🧪 Diagnóstico técnico")

        st.markdown("#### Archivos")
        st.dataframe(diag_df, use_container_width=True, hide_index=True)

        st.markdown("#### Detección de naturaleza")
        diag_cols = [
            "sistema_origen", "empresa", "archivo", "meta_codigo",
            "meta_nombre", "naturaleza", "naturaleza_confianza",
            "esperado_deudora", "error_deudora",
            "esperado_acreedora", "error_acreedora", "saldo_final_aux",
            "naturaleza_secuencia", "n_errores_saldo_secuencia",
            "max_error_saldo_secuencia"
        ]
        diag_cols = [c for c in diag_cols if c in df_audit.columns]
        st.dataframe(
            df_audit[diag_cols],
            use_container_width=True,
            hide_index=True,
        )

        st.markdown("#### Definiciones importantes")
        st.info(
            "• 'Sin referencia' significa Referencia/Docto. vacío y sin folio "
            "recuperable del Concepto.\n\n"
            "• 'Referencia libre' significa que sí existe texto en Referencia, "
            "pero no tiene forma de folio documental.\n\n"
            "• 'Antigüedad observada' no equivale a vencimiento.\n\n"
            "• 'Posible duplicado exacto' es un indicador, no una eliminación automática."
        )

    # --------------------------------------------------------------------------
    # Exportación completa
    # --------------------------------------------------------------------------
    st.divider()
    st.subheader("⬇️ Exportación")

    export_tables = {
        "Semaforo": df_audit,
        "Folios": folios,
        "Movimientos": movs,
        "Cruces_folio": df_cruces_ref,
        "Cruces_evidencia": df_evidencia,
        "Diagnostico": diag_df,
    }
    st.download_button(
        "⬇️ Descargar auditoría completa (Excel)",
        data=to_excel_workbook(export_tables),
        file_name="auditoria_master_saldos.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    if n_partidas_conciliadas or n_partidas_revisar:
        archivos_origen = [
            (uf.name, uf.getvalue()) for uf in uploaded_files
        ]
        data_marcada, nombre_marcado, mime_marcado = (
            construir_descarga_auxiliares_marcados(archivos_origen, movs)
        )
        st.download_button(
            "🎨 Descargar auxiliar(es) con conciliación marcada",
            data=data_marcada,
            file_name=nombre_marcado,
            mime=mime_marcado,
            help=(
                "Agrega una columna Conciliación y colorea las filas sin modificar "
                "póliza, fecha, documento, concepto, cargos, abonos ni saldo."
            ),
        )
    else:
        st.info(
            "No hay partidas de conciliación para marcar con los archivos cargados. "
            "Los cruces requieren movimientos relacionados entre cuentas."
        )


if __name__ == "__main__":
    main()
